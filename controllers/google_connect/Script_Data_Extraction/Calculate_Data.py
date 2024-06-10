from openpyxl import load_workbook
import pymongo
from datetime import datetime
import tempfile
from controllers.google_connect.google_file_recon import data_getter,data_getter_by_month, authenticate
import os
from database import get_client
from controllers.google_connect.Script_Data_Extraction.utils.process_store_utils import get_unique_collection_name

client = get_client()

def sumar_valores_no_nulos(lista):
    print(lista)
    suma = 0
    if isinstance(lista, int):  # Verificar si lista es un entero
        lista = [lista]  # Convertir el entero en una lista de un solo elemento
    for valor in lista:
        if valor is not None:
            # Convertir el valor a entero si es una cadena de texto
            if isinstance(valor, str):
                try:
                    valor = int(valor)
                except ValueError:
                    print(f"Error: No se pudo convertir '{valor}' a entero.")
                    continue  # Saltar este valor y continuar con el siguiente
            suma += valor
    return suma

def obtener_valores_rango_a(sheet, rango):
    valores = []
    for fila in sheet[rango]:
        for celda in fila:
            valores.append(celda.value)
    return valores

def obtener_valores_rango(sheet, rango):
    #print(sheet,rango)
    valores = []
    for fila in sheet[rango]:
        for celda in fila:
            #print(celda.value)
            valores.append(celda.value)
    # Sumar los valores no nulos
    #print(valores)
    #print("-----------------------")
    suma = sumar_valores_no_nulos(valores)
    return suma
#parsear_a_entero (p)
def p(valor_celda):
    if valor_celda is not None:
        try:
            return int(valor_celda)
        except ValueError:
            print("El valor de la celda no es un número entero")
            return 0
    else:
        return 0
    
def buscar_valor1(campo):
  # Conecta a tu base de datos MongoDB
    db = client["Data_Loss"]  # Nombre de tu base de datos
    collection = db["single_values"]  # Nombre de tu colección
    #print(campo)
    # Realiza una consulta para encontrar el valor correspondiente al campo
    resultado = collection.find_one({"nombre": campo})
    print(resultado,'/n')
    valor = resultado.get("valor")
    if valor is not None:
        return campo, valor
    else:
        valor
        return campo, 0
def buscar_valor(campo):
    # Conecta a tu base de datos MongoDB
    db = client["Data_Loss"]  # Nombre de tu base de datos
    collection = db["single_values"]  # Nombre de tu colección
    
    # Realiza una consulta para encontrar el valor correspondiente al campo
    resultado = collection.find_one({"nombre": campo})
    
    if not resultado:
        return campo, 0  # o algún valor predeterminado
    
    # Obtener la fecha de hoy
    current_date = datetime.now().strftime('%Y-%m-%d')
    current_date_obj = datetime.strptime(current_date, '%Y-%m-%d').date()
    
    # Verificar el valor actual
    fecha_vencimiento_actual = resultado.get("Fecha")
    valor_actual = resultado.get("Value")
    
    if fecha_vencimiento_actual:
        fecha_vencimiento_actual = datetime.strptime(fecha_vencimiento_actual, "%Y-%m-%d").date()
        if fecha_vencimiento_actual >= current_date_obj:
            return campo, valor_actual
    
    # Si el valor actual está vencido o no hay fecha de vencimiento, verificar en las versiones
    versiones = resultado.get("versions", [])
    
    # Ordenar las versiones por fecha de actualización más reciente
    versiones.sort(key=lambda x: x["updated_at"], reverse=True)
    
    for version in versiones:
        fecha_vencimiento = version.get("Fecha")
        if fecha_vencimiento:
            fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%Y-%m-%d").date()
            if fecha_vencimiento >= current_date_obj:
                return campo, version.get("Value")
    
    # Si no se encuentra un valor válido, devuelve un valor predeterminado o un indicador
    return campo, 0  # o algún valor predeterminado

def obtener_valores_rango_num(sheet, rango, skip):
    return sumar_valores_no_nulos(obtener_valores_rango_a(sheet,rango)[::2])


def calcular_indicador1(json_data, excel_sheets):
    print(json_data["kpi_Name"])
    # Crear un diccionario para mapear los nombres de las hojas con sus objetos de hoja de Excel
    sheet_mapping = {sheet_name: sheet for sheet_name, sheet in excel_sheets.items()}

    skip = False
    # Iterar sobre las fórmulas y calcular los valores
    resultados = []
    for formula_data in json_data["formula"]:
        #print(formula_data)
        # Construir la expresión matemática dinámicamente
        formula_expr = formula_data["formula"]
        #print(formula_expr)
        numerators = {}
        denominators = {}

        # Obtener los valores de los numeradores
        for num in formula_data["numerator"]:
            sheet_name = num["source"]
            print("SHEET NAME CALCULATES ===>>>>> ",sheet_name)
            cell_range = num["key"]
            if  num["Name"] == "EXAMENES_MEDICINA_PREVENTIVA_EMP_POBLACION_MASCULINA_20_64_ANHOS" or num["Name"] == "EXAMENES_MEDICINA_PREVENTIVA_EMP_POBLACION_FEMENINA_45_64_ANHOS":
              skip = True
            if sheet_name is None or cell_range is None:
              if num["Name"] is not None:
                valor_faltante = buscar_valor(num["Name"])
                numerators[num["Name"]] = valor_faltante[1]
            else:
              cell_range = num["key"]
              if ":" in cell_range:
                  if skip:
                    print("skip true")
                    numerator_values = obtener_valores_rango_num(sheet_mapping[sheet_name], cell_range, skip)
                  else:
                    print("skip false")
                    numerator_values = obtener_valores_rango(sheet_mapping[sheet_name], cell_range)
                  #numerator_values = obtener_valores_rango(sheet_mapping[sheet_name], cell_range)
              else:
                  numerator_values = p(sheet_mapping[sheet_name][cell_range].value)
                  #print(f"Hoja: {sheet_name}, Rango: {cell_range}, Valores: {numerator_values}")

              numerators[num["Name"]] = sumar_valores_no_nulos(numerator_values)

        # Obtener los valores de los denominadores
        for denom in formula_data["denominator"]:
            sheet_name = denom["source"]
            cell_range = denom["key"]
            if sheet_name is None or cell_range is None:
              if denom["Name"] is not None:
                valor_faltante = buscar_valor(denom["Name"])
                denominators[denom["Name"]] = valor_faltante[1]
            else:
              cell_range = denom["key"]
              if ":" in cell_range:  # Verificar si es un rango de celdas
                  denominator_values = obtener_valores_rango(sheet_mapping[sheet_name], cell_range)
              else:
                  denominator_values = p(sheet_mapping[sheet_name][cell_range].value)
                  #print(f"Hoja: {sheet_name}, Rango: {cell_range}, Valores: {denominator_values}")

              denominators[denom["Name"]] = sumar_valores_no_nulos(denominator_values)
        print(denominators)
        print(numerators)
        if not denominators:  # Si no hay valores en el denominador
            if numerators:
                # Reemplazar en la fórmula solo los valores del numerador
                for name, value in numerators.items():
                    formula_expr = formula_expr.replace(name, str(value))

                try:
                    print(formula_expr)
                    result = eval(formula_expr) * 100
                    print(result, "%", "Calculo", formula_expr)
                except ZeroDivisionError:
                    print("División por 0 en la fórmula:", formula_expr)
                    result = "DIVISION POR CERO"
            else:
                print("No hay valores para calcular el resultado.")
                result = "NO HAY VALORES"
        else:
            # Si hay valores en el denominador, realizar el cálculo normalmente
            # Reemplazar en la fórmula tanto los valores del numerador como del denominador
            for name, value in numerators.items():
                formula_expr = formula_expr.replace(name, str(value))
            for name, value in denominators.items():
                formula_expr = formula_expr.replace(name, str(value))

            try:
                print(formula_expr)
                result = eval(formula_expr) * 100
                print(result, "%", "Calculo", formula_expr)
            except ZeroDivisionError:
                print("División por 0 en la fórmula:", formula_expr)
                result = "DIVISION POR CERO"
        # Crear el nuevo JSON
        if json_data["program"] in ["Metas Sanitarias Ley 18", "Metas Sanitarias Ley 19"]:
            componente_programa = ""
            peso_relativo = 0# No incluir componente/programa si es Metas Sanitarias Ley 18 o 19
        else:
            componente_programa = json_data["component"]
            peso_relativo= formula_data["weight_relative"]
        # Crear el nuevo JSON con los resultados calculados y otras
        new_json = {
            "id": json_data["kpi_number"],
            "nombre": json_data["kpi_Name"],
            "indicador_perteneciente": json_data["program"],
            "Componente/programa": componente_programa,
            "Valor": result,
            "peso_relativo": peso_relativo,
            "peso_especifico": formula_data["weight"],  # No hay información suficiente en el JSON proporcionado para esto
            "Verification_file": json_data["Verification_file"],  # No hay información suficiente en el JSON proporcionado para esto
            "fecha_de_calculo": datetime.now()
        }

        resultados.append(new_json)
    print("SHEET_MAPPING ",sheet_mapping)
    return resultados

def procesar_json(db, collection_name, excel_sheets, db_name, collection_insert_name, recalculated=False, recalculation_month=None):
    print(db, collection_name, excel_sheets)
    # Conexión a la base de datos
    try:
        #client = pymongo.MongoClient("mongodb+srv://felipe:198252021298@cluster0.sa4li7n.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0&tlsAllowInvalidCertificates=true")
        client.server_info()  # Force connection on a request as the
                          # connect=True parameter of MongoClient seems
                          # to be useless here 
        db_collection = client[db][collection_name]
        #Verify if the collection name already exists
        new_collection_name = get_unique_collection_name(client, db_name, collection_insert_name,recalculated,recalculation_month)
        if new_collection_name != collection_insert_name:
            print(f"¡Atención! Ya existe una colección con el nombre {collection_insert_name}. Se utilizará el nombre {new_collection_name} en su lugar.")
            collection_insert_name = new_collection_name
    except pymongo.errors.ServerSelectionTimeoutError as err:
    # print the error if connection fails
        print("pymongo ERROR:", err)
    # Consulta los documentos que necesitamos
    documentos = db_collection.find({})

    # Lista para almacenar los resultados
    resultados = []

    # Itera sobre los documentos y procesa cada uno
    for documento in documentos:
        # Suponiendo que cada documento es un JSON similar a los que has estado utilizando
        resultados_json = calcular_indicador1(documento, excel_sheets)

        # Agrega los resultados a la lista
        resultados.extend(resultados_json)

    # Inserta los resultados en la colección especificada
    praps = client[db_name][collection_insert_name]
    praps.insert_many(resultados)
    return resultados
def get_data_sources():
    db = client['Data_Origin']
    collection = db['Files']
    
    data_sources = list(collection.find({}))
    return data_sources

# Crear un directorio temporal para almacenar los archivos
temp_dir = tempfile.mkdtemp()
def CalculateDataDinamic():
    try:
        service = authenticate()
    except Exception as e:
        print("Error al autenticar ")
        print(e)
        return False  # Indicate error on authentication failure
    c=0
    print("DATA CALCULATION")

    # Get data sources from the database (assuming get_data_sources is defined elsewhere)
    data_sources = get_data_sources()
    print(data_sources)

    excel_sheets = {}
    temp_files = []  # List to track temporary files for cleanup

    try:
        # Process each data source
        for data_source in data_sources:
            file_name = data_source['name']
            folder_id = data_source['folder_id']
            sheets = data_source['sheets']

            print(f"Processing data source: {file_name}")
            print(f"Folder ID: {folder_id}")
            print(f"Sheets: {sheets}")

            # Create a temporary file with a unique name using tempfile.NamedTemporaryFile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file_path = temp_file.name
                temp_files.append(temp_file_path)  # Add to list for cleanup

                print("Downloading data...")
                temp_file.write(data_getter(file_name, service, folder_id))

                print("Loading workbook...")
                workbook = load_workbook(filename=temp_file_path)
                print(temp_file_path)
                print(workbook.sheetnames)

                # Create excel_sheets dictionary with filtered sheets
                filtered_sheets = {sheet: workbook[sheet] for sheet in sheets}
                excel_sheets.update(filtered_sheets)  # Update main dictionary
                print(excel_sheets)

        # Use excel_sheets for calculations (assuming 'calcular_indicador1' is defined elsewhere)
        # ...
        print("EXCEL SHEET ======>>> ", excel_sheets)
        result = procesar_json("Clave_referencia", "Praps_Primario", excel_sheets, "KPI_PRAPS","Praps")
        if result:
            print("Procesamiento exitoso. Se han insertado datos en la colección 'Praps'.")
            c+=1
        else:
            print("Hubo un problema al procesar los datos.")
        print(result)
        print("----------------------------------------")
        MetasSanitarias = procesar_json("Clave_referencia", "Metas_Sanitarias_18", excel_sheets, "KPI_METAS_SANITARIAS_18","Metas_Sanitarias_18")
        
        if MetasSanitarias:
            print("Procesamiento exitoso. Se han insertado datos en la colección 'Metas_Sanitarias_18'.")
            c+=1
        else:
            print("Hubo un problema al procesar los datos.")
        print(MetasSanitarias)
        print("----------------------------------------")
        MetasSanitarias19 = procesar_json("Clave_referencia", "Metas_Sanitarias_19", excel_sheets, "KPI_METAS_SANITARIAS_19","Metas_Sanitarias_19")
        if MetasSanitarias19:
            print("Procesamiento exitoso. Se han insertado datos en la colección 'Metas_Sanitarias_19'.")
            c+=1
        else:
            print("Hubo un problema al procesar los datos.")    
        print("----------------------------------------")
        print(MetasSanitarias19)


        # Close temporary files (not strictly necessary, as they are automatically deleted due to delete=False)
        for temp_file_path in temp_files:
            os.close(os.open(temp_file_path, os.O_RDONLY))  # Close safely
        try:
            for temp_file_path in temp_files:
                os.remove(temp_file_path)
        except OSError as e:
            print(f"Error deleting temporary file: {temp_file_path}", e)

        if c == 3:
            return True
        else:
            return False # Indicate success

    except Exception as e:
        print("Error during data processing:")
        print(e)
        # Optionally, attempt to clean up temporary files here if needed
        
def CalculateDataDinamic_previousPeriod(month,year):
    if year == '':
        year_actual = datetime.now().year
    else:
        year_actual = year
    try:
        service = authenticate()
    except Exception as e:
        print("Error al autenticar ")
        print(e)
        return False  # Indicate error on authentication failure
    c=0
    print("DATA CALCULATION")

    # Get data sources from the database (assuming get_data_sources is defined elsewhere)
    data_sources = get_data_sources()
    print(data_sources)

    excel_sheets = {}
    temp_files = []  # List to track temporary files for cleanup

    try:
        # Process each data source
        for data_source in data_sources:
            file_name = data_source['name']
            folder_id = data_source['folder_id']
            sheets = data_source['sheets']

            print(f"Processing data source: {file_name}")
            print(f"Folder ID: {folder_id}")
            print(f"Sheets: {sheets}")

            # Create a temporary file with a unique name using tempfile.NamedTemporaryFile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file_path = temp_file.name
                temp_files.append(temp_file_path)  # Add to list for cleanup

                print("Downloading data...")
                temp_file.write(data_getter_by_month(file_name, service, folder_id,month,year_actual))

                print("Loading workbook...")
                workbook = load_workbook(filename=temp_file_path)
                print(temp_file_path)
                print(workbook.sheetnames)

                # Create excel_sheets dictionary with filtered sheets
                filtered_sheets = {sheet: workbook[sheet] for sheet in sheets}
                excel_sheets.update(filtered_sheets)  # Update main dictionary
                print(excel_sheets)

        # Use excel_sheets for calculations (assuming 'calcular_indicador1' is defined elsewhere)
        # ...
        recalculation_month = f"{year}-{int(month):02d}"

        result = procesar_json("Clave_referencia", "Praps_Primario", excel_sheets, "KPI_PRAPS","Praps", recalculated=True, recalculation_month=recalculation_month)
        if result:
            print("Procesamiento exitoso. Se han insertado datos en la colección 'Praps'.")
            c+=1
        else:
            print("Hubo un problema al procesar los datos.")
        print(result)
        print("----------------------------------------")
        MetasSanitarias = procesar_json("Clave_referencia", "Metas_Sanitarias_18", excel_sheets, "KPI_METAS_SANITARIAS_18","Metas_Sanitarias_18", recalculated=True, recalculation_month=recalculation_month)
        
        if MetasSanitarias:
            print("Procesamiento exitoso. Se han insertado datos en la colección 'Metas_Sanitarias_18'.")
            c+=1
        else:
            print("Hubo un problema al procesar los datos.")
        print(MetasSanitarias)
        print("----------------------------------------")
        MetasSanitarias19 = procesar_json("Clave_referencia", "Metas_Sanitarias_19", excel_sheets, "KPI_METAS_SANITARIAS_19","Metas_Sanitarias_19", recalculated=True, recalculation_month=recalculation_month)
        if MetasSanitarias19:
            print("Procesamiento exitoso. Se han insertado datos en la colección 'Metas_Sanitarias_19'.")
            c+=1
        else:
            print("Hubo un problema al procesar los datos.")    
        print("----------------------------------------")
        print(MetasSanitarias19)


        # Close temporary files (not strictly necessary, as they are automatically deleted due to delete=False)
        for temp_file_path in temp_files:
            os.close(os.open(temp_file_path, os.O_RDONLY))  # Close safely
        try:
            for temp_file_path in temp_files:
                os.remove(temp_file_path)
        except OSError as e:
            print(f"Error deleting temporary file: {temp_file_path}", e)

        if c == 3:
            return True
        else:
            return False # Indicate success

    except Exception as e:
        print("Error during data processing:")
        print(e)